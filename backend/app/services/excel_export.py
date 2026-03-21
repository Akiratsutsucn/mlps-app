from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


RESULT_COLORS = {
    "符合": "92D050",
    "不符合": "FF0000",
    "部分符合": "FF69B4",
    "不适用": "4472C4",
}

RISK_COLORS = {
    "高危": "FF0000",
    "中危": "FF8C00",
    "低危": "4472C4",
}

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)


def _style_header(ws, row, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _style_cell(cell, wrap=True):
    cell.alignment = Alignment(vertical="center", wrap_text=wrap)
    cell.border = THIN_BORDER


def generate_excel(project, records_by_object, issues):
    wb = Workbook()

    # Sheet 1: 被测系统基本信息
    ws1 = wb.active
    ws1.title = "被测系统基本信息"
    info_data = [
        ["系统名称", project.name],
        ["等保级别", project.security_level],
        ["被测单位", project.organization],
        ["评测日期", project.eval_date or ""],
        ["评测复核员", project.reviewer or ""],
    ]
    ws1.column_dimensions["A"].width = 20
    ws1.column_dimensions["B"].width = 50
    for r, (label, value) in enumerate(info_data, 1):
        c1 = ws1.cell(row=r, column=1, value=label)
        c1.font = Font(bold=True)
        _style_cell(c1)
        c2 = ws1.cell(row=r, column=2, value=value)
        _style_cell(c2)

    # 评测对象列表
    ws1.cell(row=len(info_data) + 2, column=1, value="评测对象列表").font = Font(bold=True, size=12)
    obj_headers = ["序号", "评测对象类型", "名称", "子类型", "备注"]
    header_row = len(info_data) + 3
    for col, h in enumerate(obj_headers, 1):
        ws1.cell(row=header_row, column=col, value=h)
    _style_header(ws1, header_row, len(obj_headers))
    for i, (obj, _) in enumerate(records_by_object, 1):
        row = header_row + i
        vals = [i, obj.object_type, obj.name, obj.sub_type or "", obj.extra_info or ""]
        for col, v in enumerate(vals, 1):
            c = ws1.cell(row=row, column=col, value=v)
            _style_cell(c)

    # Sheet 2: 检查项记录汇总
    ws2 = wb.create_sheet("检查项记录汇总")
    headers2 = ["序号", "评测对象", "对象类型", "测评单元编码", "分类", "检查内容", "检查结果", "描述"]
    for col, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=col, value=h)
    _style_header(ws2, 1, len(headers2))
    ws2.column_dimensions["F"].width = 50
    ws2.column_dimensions["H"].width = 40

    row_num = 2
    seq = 1
    for obj, obj_records in records_by_object:
        for rec in obj_records:
            item = rec.check_item
            if not item:
                continue
            vals = [
                seq, obj.name, obj.object_type,
                item.item_code or "", item.category,
                item.content, rec.result or "未填写", rec.description or ""
            ]
            for col, v in enumerate(vals, 1):
                c = ws2.cell(row=row_num, column=col, value=v)
                _style_cell(c)
                if col == 7 and rec.result in RESULT_COLORS:
                    c.fill = PatternFill(start_color=RESULT_COLORS[rec.result],
                                         end_color=RESULT_COLORS[rec.result], fill_type="solid")
                    if rec.result in ("不符合", "符合"):
                        c.font = Font(color="FFFFFF")
            row_num += 1
            seq += 1

    # Sheet 3: 问题清单
    ws3 = wb.create_sheet("问题清单")
    headers3 = ["序号", "问题描述", "风险等级", "整改建议", "甲方确认意见"]
    for col, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=col, value=h)
    _style_header(ws3, 1, len(headers3))
    ws3.column_dimensions["B"].width = 50
    ws3.column_dimensions["D"].width = 40
    ws3.column_dimensions["E"].width = 30

    for i, issue in enumerate(issues, 1):
        row = i + 1
        vals = [i, issue.description, issue.risk_level, issue.suggestion or "", issue.client_opinion or ""]
        for col, v in enumerate(vals, 1):
            c = ws3.cell(row=row, column=col, value=v)
            _style_cell(c)
            if col == 3 and issue.risk_level in RISK_COLORS:
                c.fill = PatternFill(start_color=RISK_COLORS[issue.risk_level],
                                     end_color=RISK_COLORS[issue.risk_level], fill_type="solid")
                c.font = Font(color="FFFFFF")

    return wb
